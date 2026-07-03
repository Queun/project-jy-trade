#!/usr/bin/env python3
"""
Inspect sample Excel files and infer likely header rows.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

import openpyxl
import xlrd


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def non_empty_count(row: list[str]) -> int:
    return sum(1 for value in row if value)


def likely_header_index(rows: list[list[str]]) -> int:
    best_index = 0
    best_score = -1
    for idx, row in enumerate(rows[:20]):
        values = [value for value in row if value]
        score = len(values)
        if any(keyword in "".join(values) for keyword in ("商品", "订货", "收货", "仓库", "商家编码", "数量")):
            score += 5
        if score > best_score:
            best_score = score
            best_index = idx
    return best_index


def trim_row(row: list[str]) -> list[str]:
    end = len(row)
    while end > 0 and not row[end - 1]:
        end -= 1
    return row[:end]


def inspect_xlsx(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        for row in ws.iter_rows(max_row=30, values_only=True):
            rows.append([clean(cell) for cell in row])
        header_idx = likely_header_index(rows) if rows else 0
        sheets.append(
            {
                "sheet": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "header_row": header_idx + 1,
                "headers": trim_row(rows[header_idx]) if rows else [],
                "sample_rows": [trim_row(row) for row in rows[header_idx + 1 : header_idx + 6]],
            }
        )
    return sheets


def inspect_xls(path: Path) -> list[dict[str, Any]]:
    book = xlrd.open_workbook(path)
    sheets = []
    for sheet in book.sheets():
        rows: list[list[str]] = []
        for r in range(min(sheet.nrows, 30)):
            rows.append([clean(sheet.cell_value(r, c)) for c in range(sheet.ncols)])
        header_idx = likely_header_index(rows) if rows else 0
        sheets.append(
            {
                "sheet": sheet.name,
                "max_row": sheet.nrows,
                "max_column": sheet.ncols,
                "header_row": header_idx + 1,
                "headers": trim_row(rows[header_idx]) if rows else [],
                "sample_rows": [trim_row(row) for row in rows[header_idx + 1 : header_idx + 6]],
            }
        )
    return sheets


def inspect(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        sheets = inspect_xlsx(path)
    elif suffix == ".xls":
        sheets = inspect_xls(path)
    else:
        raise ValueError(f"Unsupported file type: {path}")
    return {"file": str(path), "sheets": sheets}


def main() -> int:
    parser = argparse.ArgumentParser(description="Inspect Excel workbook schemas")
    parser.add_argument("files", nargs="+")
    args = parser.parse_args()

    result = [inspect(Path(file)) for file in args.files]
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
