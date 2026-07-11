from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


output_path = Path("outputs/fixtures/确定单-自动分配测试.xlsx")
headers = [
    "审批单号", "通知单号", "收货地编码", "收货地名称", "业务员", "截止日期", "商品编码", "商品条码",
    "商品名称", "规格", "订货数量", "实际发货数量", "合同进价", "主仓", "临期仓",
]
rows = [
    ["MOCK-APP-001", "MOCK-FULL-001", "207159", "Ole太原王府井店", "测试业务员", "2026-07-31", "6941594584092", "6941594584092", "【中小样】兰蔻全新清滢保湿柔肤水旅行装", "测试规格", 8, 5, 12.5, 999, 999],
    ["MOCK-APP-002", "MOCK-FAIR-001", "207086", "Ole长春万象城店", "测试业务员", "2026-07-31", "747930061007", "747930061007", "海蓝之谜经典精华面霜", "测试规格", 8, 4, 20, 999, 999],
    ["MOCK-APP-003", "MOCK-FAIR-002", "205675", "Ole石家庄万象城店", "测试业务员", "2026-07-31", "747930061007", "747930061007", "海蓝之谜经典精华面霜", "测试规格", 8, 4, 20, 999, 999],
    ["MOCK-APP-004", "MOCK-BLOCKED-001", "205677", "Ole太原万象城店", "测试业务员", "2026-07-31", "6974354840411", "6974354840411", "【中小样】养生堂专研润泽修护精华乳", "测试规格", 6, 3, 15, 999, 999],
    ["MOCK-APP-005", "MOCK-ZERO-001", "205686", "Ole济南万象城店", "测试业务员", "2026-07-31", "6941594584092", "6941594584092", "【中小样】兰蔻全新清滢保湿柔肤水旅行装", "测试规格", 5, 0, 12.5, 999, 999],
]

workbook = Workbook()
sheet = workbook.active
sheet.title = "确定单"
sheet.append(headers)
for row in rows:
    sheet.append(row)

header_fill = PatternFill("solid", fgColor="DCE6F1")
for cell in sheet[1]:
    cell.font = Font(name="Arial", size=10, bold=True, color="1F2937")
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
for row in sheet.iter_rows(min_row=2):
    for cell in row:
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center")

widths = [16, 20, 14, 24, 14, 14, 18, 18, 42, 14, 12, 16, 12, 10, 10]
for column, width in zip(sheet.columns, widths):
    sheet.column_dimensions[column[0].column_letter].width = width
sheet.freeze_panes = "A2"
sheet.auto_filter.ref = f"A1:O{len(rows) + 1}"
sheet.row_dimensions[1].height = 24

output_path.parent.mkdir(parents=True, exist_ok=True)
workbook.save(output_path)
print(output_path)
